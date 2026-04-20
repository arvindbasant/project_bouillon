"""
build_excel.py  —  produce the Week-1 Excel deliverable.

Tabs:
  1. Summary          — site × channel pivot (cases + tonnes)
  2. SKU Master       — full 57-row SKU table (clean headers, formatted)
  3. Monthly Volume   — long-format monthly fact table
  4. Format × Site    — format heatmap in tonnes
  5. S1 vs S2         — capability gap breakdown
  6. Data Notes       — provenance / assumptions

Run from repo root:  python3 deliverable/build_excel.py
Writes:  deliverable/Bouillon_Week1_Data.xlsx
"""

import csv
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers as xl_numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl not installed. Run: pip install openpyxl")

ROOT    = Path(__file__).parent.parent
DATA    = ROOT / "data"
OUT     = Path(__file__).parent / "Bouillon_Week1_Data.xlsx"

# ── Palette ──────────────────────────────────────────────────────────────────
LERMA_BG   = "2F5A8A"
INDY_BG    = "2E8B57"
BATAVIA_BG = "C97B1F"
HDR_BG     = "1F3864"
HDR2_BG    = "2F5597"
SUB_BG     = "D6E4F0"
S1_BG      = "C6EFCE"
S2_BG      = "FFCCCC"
ATINDY_BG  = "BDD7EE"
WHITE_FG   = "FFFFFF"
DARK_FG    = "1F1F1F"
NOTE_BG    = "FFF2CC"

def hdr(text, bold=True, fg=WHITE_FG, bg=HDR_BG, size=11, wrap=False):
    f = Font(name="Calibri", bold=bold, color=fg, size=size)
    fil = PatternFill("solid", fgColor=bg)
    aln = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return f, fil, aln

def apply(cell, font=None, fill=None, align=None, fmt=None, border=None):
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if fmt:    cell.number_format = fmt
    if border: cell.border    = border

def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        length = max(
            (len(str(cell.value)) if cell.value else 0 for cell in col),
            default=8
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


# ── Load data ─────────────────────────────────────────────────────────────────
def load_sku():
    rows = []
    with open(DATA / "sku_master.csv") as f:
        for r in csv.DictReader(f):
            r["total_cases_2025"]  = float(r["total_cases_2025"]  or 0)
            r["total_tonnes_2025"] = float(r["total_tonnes_2025"] or 0)
            r["country"]           = (r["country"] or "").strip()
            rows.append(r)
    return rows

def load_monthly():
    rows = []
    with open(DATA / "monthly_volume.csv") as f:
        for r in csv.DictReader(f):
            r["cases"]  = float(r["cases"]  or 0)
            r["tonnes"] = float(r["tonnes"] or 0)
            rows.append(r)
    return rows


# ── Tab 1: Summary ────────────────────────────────────────────────────────────
def tab_summary(wb, sku):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    sites    = ["Lerma", "Indy", "Batavia", "TOTAL"]
    channels = ["US Retail", "UFS", "CDA Retail", "UI", "TOTAL"]
    site_bg  = {
        "Lerma":   LERMA_BG,
        "Indy":    INDY_BG,
        "Batavia": BATAVIA_BG,
        "TOTAL":   HDR_BG,
    }

    # Aggregate
    mat_c = defaultdict(lambda: defaultdict(float))
    mat_t = defaultdict(lambda: defaultdict(float))
    for r in sku:
        su, cn = r["sourcing_unit"], r["country"]
        mat_c[su][cn] += r["total_cases_2025"]
        mat_t[su][cn] += r["total_tonnes_2025"]
    for su in ["Lerma", "Indy", "Batavia"]:
        for cn in channels[:-1]:
            mat_c["TOTAL"][cn] += mat_c[su][cn]
            mat_t["TOTAL"][cn] += mat_t[su][cn]
    for su in sites:
        mat_c[su]["TOTAL"] = sum(mat_c[su][cn] for cn in channels[:-1])
        mat_t[su]["TOTAL"] = sum(mat_t[su][cn] for cn in channels[:-1])
    net_c = mat_c["TOTAL"]["TOTAL"]
    net_t = mat_t["TOTAL"]["TOTAL"]

    # Title
    ws.merge_cells("A1:K1")
    ws["A1"] = "NA Bouillon — 2025 Baseline  |  Site × Channel Summary"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color=WHITE_FG)
    ws["A1"].fill      = PatternFill("solid", fgColor=HDR_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells("A2:K2")
    ws["A2"] = "Source: RT and UFS Bouillon 4.13.xlsx  |  Tonnes = Cases × Tonnes/CS (directional, not mass-balance)"
    ws["A2"].font      = Font(name="Calibri", size=10, italic=True, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers (cases block + tonnes block)
    ws.merge_cells("B4:F4")
    ws["B4"] = "CASES (2025)"
    ws["B4"].font = Font(bold=True, color=WHITE_FG, size=11)
    ws["B4"].fill = PatternFill("solid", fgColor=HDR2_BG)
    ws["B4"].alignment = Alignment(horizontal="center")

    ws.merge_cells("G4:K4")
    ws["G4"] = "TONNES (2025, directional)"
    ws["G4"].font = Font(bold=True, color=WHITE_FG, size=11)
    ws["G4"].fill = PatternFill("solid", fgColor=HDR2_BG)
    ws["G4"].alignment = Alignment(horizontal="center")

    # Channel sub-headers
    ws["A5"] = "Site"
    for j, ch in enumerate(channels):
        for blk in [1, 6]:  # cases offset=1, tonnes offset=6
            c = ws.cell(row=5, column=blk + j, value=ch)
            c.font = Font(bold=True, color=WHITE_FG, size=10)
            c.fill = PatternFill("solid", fgColor=HDR_BG)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws["A5"].font = Font(bold=True, color=WHITE_FG, size=10)
    ws["A5"].fill = PatternFill("solid", fgColor=HDR_BG)
    ws.row_dimensions[5].height = 30

    # Data rows
    row = 6
    for su in sites:
        ws.cell(row=row, column=1, value=su).font = Font(
            bold=True, color=WHITE_FG if su != "TOTAL" else WHITE_FG, size=11
        )
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=site_bg[su])
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")

        is_total_row = (su == "TOTAL")
        for j, ch in enumerate(channels):
            is_total_col = (ch == "TOTAL")

            # Cases
            cv = mat_c[su][ch]
            cc = ws.cell(row=row, column=2 + j, value=cv if cv else None)
            cc.number_format = '#,##0'
            cc.alignment = Alignment(horizontal="right")
            if is_total_row or is_total_col:
                cc.font = Font(bold=True, size=11)
                cc.fill = PatternFill("solid", fgColor="DDEEFF")
            if is_total_row and is_total_col:
                cc.font = Font(bold=True, size=12, color=WHITE_FG)
                cc.fill = PatternFill("solid", fgColor=HDR_BG)

            # Tonnes
            tv = mat_t[su][ch]
            tc = ws.cell(row=row, column=7 + j, value=round(tv, 1) if tv else None)
            tc.number_format = '#,##0.0'
            tc.alignment = Alignment(horizontal="right")
            if is_total_row or is_total_col:
                tc.font = Font(bold=True, size=11)
                tc.fill = PatternFill("solid", fgColor="DDEEFF")
            if is_total_row and is_total_col:
                tc.font = Font(bold=True, size=12, color=WHITE_FG)
                tc.fill = PatternFill("solid", fgColor=HDR_BG)

        ws.row_dimensions[row].height = 22
        row += 1

    # Share rows (% of network tonnes)
    ws.cell(row=row, column=1, value="% of net tonnes").font = Font(italic=True, size=10)
    for j, su in enumerate(sites):
        tv = mat_t[su]["TOTAL"]
        c = ws.cell(row=row, column=7 + j, value=tv / net_t if net_t else 0)
        c.number_format = "0%"
        c.font = Font(italic=True, size=10)
        c.alignment = Alignment(horizontal="right")
    row += 1
    ws.cell(row=row, column=1, value="% of net cases").font = Font(italic=True, size=10)
    for j, su in enumerate(sites):
        cv = mat_c[su]["TOTAL"]
        c = ws.cell(row=row, column=2 + j, value=cv / net_c if net_c else 0)
        c.number_format = "0%"
        c.font = Font(italic=True, size=10)
        c.alignment = Alignment(horizontal="right")

    auto_width(ws)
    ws.column_dimensions["A"].width = 14
    ws.freeze_panes = "B6"
    print("  Tab: Summary")


# ── Tab 2: SKU Master ─────────────────────────────────────────────────────────
def tab_sku_master(wb, sku):
    ws = wb.create_sheet("SKU Master")
    ws.sheet_view.showGridLines = False

    COLS = [
        ("Row",              "row",                 "#,##0",    8),
        ("Site",             "sourcing_unit",        "@",        10),
        ("DU",               "du",                  "@",        12),
        ("Brand",            "brand",               "@",        10),
        ("Sub-Category",     "sub_category",        "@",        18),
        ("Technology",       "technology_detailed", "@",        22),
        ("Line",             "line",                "@",        10),
        ("Size",             "size",                "@",        8),
        ("Channel",          "country",             "@",        12),
        ("Sales Org",        "sales_org_opco",      "@",        12),
        ("Product DU",       "product_du",          "@",        36),
        ("Tonnes/CS",        "pouches_cs",          "0.000000", 12),
        ("2025 Cases",       "total_cases_2025",    "#,##0.0",  12),
        ("2025 Tonnes",      "total_tonnes_2025",   "#,##0.0",  12),
    ]

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws["A1"] = "NA Bouillon — 2025 SKU Master"
    ws["A1"].font = Font(bold=True, size=13, color=WHITE_FG)
    ws["A1"].fill = PatternFill("solid", fgColor=HDR_BG)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    # Headers
    site_bg = {"Lerma": LERMA_BG, "Indy": INDY_BG, "Batavia": BATAVIA_BG}
    for j, (label, _, fmt, w) in enumerate(COLS):
        c = ws.cell(row=2, column=j + 1, value=label)
        c.font = Font(bold=True, color=WHITE_FG, size=10)
        c.fill = PatternFill("solid", fgColor=HDR2_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j + 1)].width = w
    ws.row_dimensions[2].height = 30

    # Data
    for i, r in enumerate(sku):
        row = 3 + i
        su_bg = site_bg.get(r["sourcing_unit"], "F0F0F0")
        for j, (_, key, fmt, _) in enumerate(COLS):
            val = r.get(key, "")
            try:
                val = float(val) if fmt not in ("@",) else val
            except (TypeError, ValueError):
                pass
            c = ws.cell(row=row, column=j + 1, value=val)
            c.number_format = fmt
            c.alignment = Alignment(vertical="center",
                                    horizontal="right" if fmt != "@" else "left")
            c.font = Font(size=10)
        # Colour the Site cell
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=su_bg)
        ws.cell(row=row, column=2).font = Font(bold=True, color=WHITE_FG, size=10)
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{2 + len(sku)}"
    print("  Tab: SKU Master")


# ── Tab 3: Monthly Volume ─────────────────────────────────────────────────────
def tab_monthly(wb, monthly):
    ws = wb.create_sheet("Monthly Volume")
    ws.sheet_view.showGridLines = False

    COLS = [
        ("DU",          "du",           "@",      12),
        ("Site",        "sourcing_unit","@",      10),
        ("Product DU",  "product_du",  "@",      36),
        ("Channel",     "country",     "@",      12),
        ("Month",       "month",       "@",       8),
        ("Cases",       "cases",       "#,##0.0", 12),
        ("Tonnes",      "tonnes",      "#,##0.0", 10),
    ]

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws["A1"] = "NA Bouillon — 2025 Monthly Volume (long format)"
    ws["A1"].font = Font(bold=True, size=13, color=WHITE_FG)
    ws["A1"].fill = PatternFill("solid", fgColor=HDR_BG)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    for j, (label, _, fmt, w) in enumerate(COLS):
        c = ws.cell(row=2, column=j + 1, value=label)
        c.font = Font(bold=True, color=WHITE_FG, size=10)
        c.fill = PatternFill("solid", fgColor=HDR2_BG)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j + 1)].width = w
    ws.row_dimensions[2].height = 28

    site_bg = {"Lerma": LERMA_BG, "Indy": INDY_BG, "Batavia": BATAVIA_BG}
    for i, r in enumerate(monthly):
        row = 3 + i
        for j, (_, key, fmt, _) in enumerate(COLS):
            val = r.get(key, "")
            try:
                val = float(val) if fmt != "@" else val
            except (TypeError, ValueError):
                pass
            c = ws.cell(row=row, column=j + 1, value=val)
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right" if fmt != "@" else "left")
            c.font = Font(size=9)
        su = r.get("sourcing_unit", "")
        if su in site_bg:
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=site_bg[su])
            ws.cell(row=row, column=2).font = Font(bold=True, color=WHITE_FG, size=9)
        ws.row_dimensions[row].height = 15

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{2 + len(monthly)}"
    print("  Tab: Monthly Volume")


# ── Tab 4: Format × Site (tonnes heatmap) ─────────────────────────────────────
def tab_format_site(wb, sku):
    ws = wb.create_sheet("Format × Site")
    ws.sheet_view.showGridLines = False

    def _chan(r): return (r["country"] or "").strip()
    format_rules = [
        ("7.9 oz",                 lambda r: r["size"] == "7.9oz"),
        ("3.5 oz",                 lambda r: r["size"] == "3.5oz"),
        ("15.9 oz",                lambda r: r["size"] == "15.9oz"),
        ("2 lb",                   lambda r: r["size"] == "2lb"),
        ("2.5 lb",                 lambda r: r["size"] == "2.5lb"),
        ("2.6 oz",                 lambda r: r["size"] == "2.6oz"),
        ("40.5 oz",                lambda r: r["size"] == "40.5oz"),
        ("4.4 lb (UFS)",           lambda r: r["size"] == "4.4lb"),
        ("7.9 lb (UFS)",           lambda r: r["size"] == "7.9lb"),
        ("25 lb bulk Caldo (UFS)", lambda r: r["size"] in ("25 lb", "0") and _chan(r) == "UFS"),
        ("160 g Zero Salt (CDA)",  lambda r: r["size"] == "0" and _chan(r) == "CDA Retail"),
    ]
    sites = ["Indy", "Lerma", "Batavia"]

    mat_t = {fmt: {s: 0.0 for s in sites} for fmt, _ in format_rules}
    mat_c = {fmt: {s: 0.0 for s in sites} for fmt, _ in format_rules}
    for r in sku:
        su = r["sourcing_unit"]
        if su not in sites: continue
        for label, rule in format_rules:
            if rule(r):
                mat_t[label][su] += r["total_tonnes_2025"]
                mat_c[label][su] += r["total_cases_2025"]
                break

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Format × Site — 2025 volume (tonnes + cases)"
    ws["A1"].font = Font(bold=True, size=13, color=WHITE_FG)
    ws["A1"].fill = PatternFill("solid", fgColor=HDR_BG)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Header row
    headers = ["Format",
               "Indy\ntonnes", "Lerma\ntonnes", "Batavia\ntonnes", "Total\ntonnes",
               "Indy\ncases",  "Lerma\ncases",  "Batavia\ncases",  "Total\ncases"]
    site_bgs = [INDY_BG, LERMA_BG, BATAVIA_BG, HDR_BG,
                INDY_BG, LERMA_BG, BATAVIA_BG, HDR_BG]
    for j, (h, bg) in enumerate(zip(headers, [""] + site_bgs)):
        c = ws.cell(row=2, column=j + 1, value=h)
        c.font = Font(bold=True, color=WHITE_FG, size=10)
        c.fill = PatternFill("solid", fgColor=(bg if bg else HDR_BG))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    # Data rows
    for i, (label, _) in enumerate(format_rules):
        row = 3 + i
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        tot_t = sum(mat_t[label].values())
        tot_c = sum(mat_c[label].values())

        at_indy = mat_t[label]["Indy"] > 0
        fill_yes = PatternFill("solid", fgColor=S1_BG)
        fill_no  = PatternFill("solid", fgColor="F5F5F5")

        for j, su in enumerate(sites):
            tv = mat_t[label][su]
            cv = mat_c[label][su]
            c_t = ws.cell(row=row, column=2 + j, value=round(tv, 1) if tv else None)
            c_t.number_format = "#,##0.0"
            c_t.alignment = Alignment(horizontal="right")
            c_t.fill = fill_yes if tv > 0 else fill_no
            if tv == 0:
                c_t.value = "—"
                c_t.alignment = Alignment(horizontal="center")
                c_t.font = Font(color="AAAAAA")

            c_c = ws.cell(row=row, column=6 + j, value=round(cv, 0) if cv else None)
            c_c.number_format = "#,##0"
            c_c.alignment = Alignment(horizontal="right")
            c_c.fill = fill_yes if cv > 0 else fill_no
            if cv == 0:
                c_c.value = "—"
                c_c.alignment = Alignment(horizontal="center")
                c_c.font = Font(color="AAAAAA")

        # Totals
        c_tt = ws.cell(row=row, column=5, value=round(tot_t, 1))
        c_tt.number_format = "#,##0.0"
        c_tt.font = Font(bold=True, size=11)
        c_tt.fill = PatternFill("solid", fgColor=SUB_BG)
        c_tt.alignment = Alignment(horizontal="right")

        c_tc = ws.cell(row=row, column=9, value=round(tot_c, 0))
        c_tc.number_format = "#,##0"
        c_tc.font = Font(bold=True, size=11)
        c_tc.fill = PatternFill("solid", fgColor=SUB_BG)
        c_tc.alignment = Alignment(horizontal="right")

        ws.row_dimensions[row].height = 20

    # Grand total row
    gr = 3 + len(format_rules)
    ws.cell(row=gr, column=1, value="TOTAL").font = Font(bold=True, size=12, color=WHITE_FG)
    ws.cell(row=gr, column=1).fill = PatternFill("solid", fgColor=HDR_BG)
    net_t = {su: sum(mat_t[lbl][su] for lbl, _ in format_rules) for su in sites}
    net_c = {su: sum(mat_c[lbl][su] for lbl, _ in format_rules) for su in sites}
    for j, su in enumerate(sites):
        for blk, d, fmt in [(2, net_t, "#,##0.0"), (6, net_c, "#,##0")]:
            c = ws.cell(row=gr, column=blk + j, value=round(d[su], 1))
            c.number_format = fmt
            c.font = Font(bold=True, color=WHITE_FG)
            c.fill = PatternFill("solid", fgColor=HDR_BG)
            c.alignment = Alignment(horizontal="right")
    for col, val, fmt in [
        (5, sum(net_t.values()), "#,##0.0"),
        (9, sum(net_c.values()), "#,##0"),
    ]:
        c = ws.cell(row=gr, column=col, value=round(val, 1))
        c.number_format = fmt
        c.font = Font(bold=True, color=WHITE_FG)
        c.fill = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="right")
    ws.row_dimensions[gr].height = 24

    col_widths = [26, 14, 14, 14, 14, 14, 14, 14, 14]
    for j, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(j + 1)].width = w
    ws.freeze_panes = "B3"
    print("  Tab: Format × Site")


# ── Tab 5: S1 vs S2 Capability Gap ───────────────────────────────────────────
def tab_capability_gap(wb, sku):
    ws = wb.create_sheet("S1 vs S2 Gap")
    ws.sheet_view.showGridLines = False

    s1_formats = {"2lb", "15.9oz"}

    rows_indy = [r for r in sku if r["sourcing_unit"] == "Indy"]
    rows_s1   = [r for r in sku if r["sourcing_unit"] == "Lerma" and r["size"] in s1_formats]
    rows_s2   = [r for r in sku if r["sourcing_unit"] != "Indy"
                 and not (r["sourcing_unit"] == "Lerma" and r["size"] in s1_formats)]

    def totals(lst):
        return sum(r["total_tonnes_2025"] for r in lst), sum(r["total_cases_2025"] for r in lst)

    indy_t, indy_c = totals(rows_indy)
    s1_t,   s1_c   = totals(rows_s1)
    s2_t,   s2_c   = totals(rows_s2)
    net_t = indy_t + s1_t + s2_t
    net_c = indy_c + s1_c + s2_c

    ws.merge_cells("A1:G1")
    ws["A1"] = "Capability Gap  —  S1 (Maximize Indy) vs S2 (Full Consolidation)"
    ws["A1"].font = Font(bold=True, size=13, color=WHITE_FG)
    ws["A1"].fill = PatternFill("solid", fgColor=HDR_BG)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    HDR = ["Segment", "Tonnes", "% Net Tonnes", "Cases", "% Net Cases",
           "What it requires", "Risk"]
    for j, h in enumerate(HDR):
        c = ws.cell(row=2, column=j + 1, value=h)
        c.font = Font(bold=True, color=WHITE_FG, size=10)
        c.fill = PatternFill("solid", fgColor=HDR2_BG)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    DATA_ROWS = [
        ("Already at Indy",
         indy_t, indy_t / net_t, indy_c, indy_c / net_c,
         "Existing Indy output — 1 line, 2 formats, Retail only",
         "Low — already running", ATINDY_BG),
        ("S1 addressable (Lerma 2 lb + 15.9 oz)",
         s1_t, s1_t / net_t, s1_c, s1_c / net_c,
         "Transfer Lerma volume that matches Indy formats. Capacity headroom question.",
         "Medium — line 15 capacity ceiling unknown", S1_BG),
        ("S2 capability build (all other volume)",
         s2_t, s2_t / net_t, s2_c, s2_c / net_c,
         "6 new formats · 4+ new technologies · UFS footprint from scratch · Capex + R&D harmonization",
         "High — capability build, not a line transfer", S2_BG),
        ("NETWORK TOTAL",
         net_t, 1.0, net_c, 1.0, "", "", HDR_BG),
    ]

    for i, (seg, t, tp, c, cp, what, risk, bg) in enumerate(DATA_ROWS):
        row = 3 + i
        vals = [seg, round(t, 0), tp, round(c, 0), cp, what, risk]
        fmts = ["@",  "#,##0",   "0%", "#,##0", "0%", "@",  "@"]
        for j, (v, f) in enumerate(zip(vals, fmts)):
            cell = ws.cell(row=row, column=j + 1, value=v)
            cell.number_format = f
            cell.alignment = Alignment(vertical="center", wrap_text=True,
                                       horizontal="right" if f != "@" else "left")
            cell.fill = PatternFill("solid", fgColor=bg)
            if i == len(DATA_ROWS) - 1:
                cell.font = Font(bold=True, color=WHITE_FG, size=11)
        ws.row_dimensions[row].height = 36

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 54
    ws.column_dimensions["G"].width = 38
    print("  Tab: S1 vs S2 Gap")


# ── Tab 6: Data Notes ─────────────────────────────────────────────────────────
def tab_data_notes(wb):
    ws = wb.create_sheet("Data Notes")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"] = "Data Provenance & Assumptions"
    ws["A1"].font = Font(bold=True, size=13, color=WHITE_FG)
    ws["A1"].fill = PatternFill("solid", fgColor=HDR_BG)
    ws["A1"].alignment = Alignment(horizontal="center")

    notes = [
        ("Source file",
         "RT and UFS Bouillon 4.13.xlsx  (same data as RT and UFS Bouillon 4.13.csv, verified identical)"),
        ("Tonnes derivation",
         "Tonnes = Cases × Tonnes/CS (column L of source workbook). "
         "Directional weight — not mass-balance reconciled against production records."),
        ("Size = 0 encoding",
         "Source workbook uses Size = 0 for two unrelated product groups: "
         "(a) 25 lb bulk Caldo UFS drums — merged with the 25 lb row (Knorr Caldo De Tomate) to form '25 lb bulk Caldo (UFS)' at 73,699 cases / 838 t; "
         "(b) 160 g Zero Salt CDA retail pouches — 11,545 cases / 22 t. "
         "These are separated in all analysis by channel (UFS vs CDA Retail)."),
        ("Multi-site SKUs",
         "9 SKUs appear at both Lerma and Indy in 2025 — these are in-flight transfers. "
         "The monthly profile (Sep-25 = 0 at Lerma, spike at Indy in Oct-25) marks the coordinated cut-over."),
        ("Negative monthly cells",
         "3 monthly cells at Indy carry negative values (-480, -192, -98). "
         "These are inventory/ERP corrections during the Lerma→Indy ramp-up. Preserved as-is."),
        ("Batavia '2 lb' volume",
         "Batavia has ~7,440 cases of size '2lb' in the source data. "
         "Treated as a residual / labelling artefact — not evidence Batavia runs the 2 lb format at scale."),
        ("Batavia line '0'",
         "Line coded as '0' at Batavia carries 45K cases. "
         "Interpreted as an unmapped or packaging-only line, not a zero-index error."),
        ("Scope",
         "NA network only: Lerma (MX), Indy (MO), Batavia (WI). "
         "Co-manufacturers not present in dataset — flagged as open data request D3."),
        ("Tonnes/CS column",
         "In sku_master.csv the column named 'pouches_cs' holds the actual Tonnes/CS values from the source "
         "(what is used to compute total_tonnes_2025). The 'tonnes_cs' column appears to be a different weight field from the source — not used in analysis."),
    ]

    for j, h in enumerate(["#", "Item", "Detail", ""]):
        c = ws.cell(row=2, column=j + 1, value=h)
        c.font = Font(bold=True, color=WHITE_FG)
        c.fill = PatternFill("solid", fgColor=HDR2_BG)
    ws.row_dimensions[2].height = 24

    for i, (item, detail) in enumerate(notes):
        row = 3 + i
        ws.cell(row=row, column=1, value=i + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=item).font = Font(bold=True, size=10)
        ws.cell(row=row, column=3, value=detail).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3).font = Font(size=10)
        ws.row_dimensions[row].height = 50
        for col in range(1, 4):
            ws.cell(row=row, column=col).fill = PatternFill(
                "solid", fgColor="FFFDE7" if i % 2 == 0 else "FFFFFF"
            )

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 90
    print("  Tab: Data Notes")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    sku     = load_sku()
    monthly = load_monthly()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    print("Building Excel tabs…")
    tab_summary(wb, sku)
    tab_sku_master(wb, sku)
    tab_monthly(wb, monthly)
    tab_format_site(wb, sku)
    tab_capability_gap(wb, sku)
    tab_data_notes(wb)

    wb.save(OUT)
    print(f"\nSaved → {OUT.relative_to(Path.cwd())}")
    print(f"  Tabs: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
